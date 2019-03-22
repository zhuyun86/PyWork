import math
import sys
import os
from openpyxl import load_workbook


def readGeneMutation():
    wb = load_workbook(r'C:\Users\12sigma\Desktop\标注 突变信息.xlsx')
    sheet = wb.get_sheet_by_name("Sheet1")
    for i in range(2, 22): #2-21
        _row = i
        sample = str(sheet[_row][1].value)
        res = ''
        for i, v in enumerate(sheet[_row]): #30-312-333
            if i not in range(30, 334):
                continue
            if v.value == '-' or sheet[1][i].value == 'Pathway':
                continue
            res += sheet[1][i].value + ': ' + str(v.value) + '\n'
        with open(sample+'_gene.txt', 'w', encoding='UTF-8') as fw:
            fw.write(res)


dict_header = {'抗病毒治疗': 'Pre-operative-antiviral', '术前肿瘤治疗情况（射频、TACE等）': 'Pre-operative-TACE', '子灶': 'Satellite', '包膜': 'Break',
        '坏死': 'Necrosis', '肝外侵犯（写明部位）': 'Extrainvasion', '镜下包膜有无突破': 'Microbreak', '镜下多灶性生长': 'MultiSatelliteTumor', '镜下子灶': 'Microscopic-satellite',
        '微血管癌栓': 'Microvascular-invasion', '肝硬化': 'Cirrhosis', '肝炎': 'Hepatitis', '肉眼癌栓': 'Vein-emboli', '分型': 'Type', '分级': 'Histological-grade',
         '术后TACE（注明治疗次数）': 'Post-operative-TACE', '术后抗病毒治疗（注明化疗药物）': 'Post-operative-antiviral'}

def readPathologicalFeatures():
    wb = load_workbook(r'C:\Users\12sigma\Desktop\病人信息登记(1).xlsx')
    sheet = wb.get_sheet_by_name("Sheet2")
    for _row in range(2, 20): #2-19
        sample = str(sheet[_row][4].value)
        res = ''
        for i, v in enumerate(sheet[_row]): #0-69
            if i not in [52, 53, 54]:
                if i not in [31, 32, 56, 57]:
                    if i not in range(39, 50):
                        continue
            # if v.value == '无':
            #     continue
            if sheet[1][i].value in dict_header.keys():
                value = v.value
                if sheet[1][i].value == '分级':
                    if '低' in v.value:
                        if '中' in v.value:
                            value = 2
                        else:
                            value = 1
                    elif '中' in v.value:
                        if '高' in v.value:
                            value = 4
                        else:
                            value = 3
                    elif '高' in v.value:
                        value = 5
                    else:
                        value = 0
                elif sheet[1][i].value == '包膜':
                    if '不' in v.value:
                        value = 1
                    else:
                        value = 0
                elif sheet[1][i].value in ['肝外侵犯（写明部位）', '抗病毒治疗', '术前肿瘤治疗情况（射频、TACE等）',
                                            '子灶', '坏死', '肝外侵犯（写明部位）', '镜下包膜有无突破', '镜下多灶性生长',
                                            '镜下子灶', '微血管癌栓', '肝硬化', '肝炎', '肉眼癌栓',
                                            '术后TACE（注明治疗次数）', '术后抗病毒治疗（注明化疗药物）']:
                    if '无' in v.value:
                        value = 0
                    else:
                        value = 1
                res += dict_header[sheet[1][i].value] + ': ' + str(value) + '\n'
            else:
                value = v.value
                if sheet[1][i].value == 'BCLC':
                    if 'A' in v.value and 'NA' != v.value:
                        value = 'A'
                    elif 'B' in v.value:
                        value = 'B'
                    elif 'C' in v.value:
                        value = 'C'
                    elif 'D' in v.value:
                        value = 'D'
                res += sheet[1][i].value + ': ' + str(value) + '\n'
            with open(sample+'_features.txt', 'w', encoding='UTF-8') as fw:
                fw.write(res)

if __name__ == '__main__':

    readPathologicalFeatures()
    readGeneMutation()
